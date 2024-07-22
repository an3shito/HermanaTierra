"""             Importación de Módulos           """

import customtkinter
import openpyxl
import mysql.connector
from mysql.connector import Error
from tkinter import ttk, filedialog, messagebox
from .vista_inventarios import AgregarInventarioVentana

class PanelInventario(customtkinter.CTkFrame):
    
    """ Panel que mostrará los inventarios a ver """
    
    def __init__(self, padre, volver_func):
        super().__init__(padre)
        
        # Configurar el grid principal 
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)  # Añadir una fila más para el botón
        self.grid_rowconfigure(2, weight=1)  # Asegurarse de tener espacio para los botones

        # Menú de opciones 
        self.combobox = customtkinter.CTkOptionMenu(self,
                                                    values=["option 1", "option 2"],
                                                    command=self.llamar_menu_opciones)
        self.combobox.grid(row=0, column=0, sticky="nw", padx=10, pady=10)
        self.combobox.set("option 1")  # Set initial value

        # Mostrar el panel
        self.mostrar_panel()
        
        # Crear el Treeview para mostrar los datos
        self.treeview = ttk.Treeview(self, columns=("NO", "NOMBRE DEL SUMINISTRO", "PRESENTACIÓN",
                                                    "LOTE", "F/V","CANTIDAD","PRECIO","TOTAL"), show='headings')
        self.treeview.grid(row=1, column=0, sticky="nsew")
        self.treeview.heading("NO", text="NO")
        self.treeview.heading("NOMBRE DEL SUMINISTRO", text="NOMBRE DEL SUMINISTRO")
        self.treeview.heading("PRESENTACIÓN", text="PRESENTACIÓN")
        self.treeview.heading("LOTE", text="LOTE")
        self.treeview.heading("F/V", text="F/V")
        self.treeview.heading("CANTIDAD", text="CANTIDAD")
        self.treeview.heading("PRECIO", text="PRECIO")
        self.treeview.heading("TOTAL", text="TOTAL")
        
        # Configurar tamaños de las columnas
        self.treeview.column("NO", width=50, anchor='center')
        self.treeview.column("NOMBRE DEL SUMINISTRO", width=200, anchor='w')
        self.treeview.column("PRESENTACIÓN", width=150, anchor='w')
        self.treeview.column("LOTE", width=80, anchor='center')
        self.treeview.column("F/V", width=80, anchor='center')
        self.treeview.column("CANTIDAD", width=100, anchor='center')
        self.treeview.column("PRECIO", width=100, anchor='center')
        self.treeview.column("TOTAL", width=100, anchor='center')
        
        # Botón para agregar inventario
        self.boton_agregar_inventario = customtkinter.CTkButton(self, text="Agregar Inventarios",
                                                                width=100, height=32,
                                                                command=self.abrir_agregar_inventario)
        self.boton_agregar_inventario.grid(row=2, column=0, sticky="sw", padx=10, pady=10)
        
        # Botón para subir inventario
        self.boton_subir_inventario = customtkinter.CTkButton(self, text="Subir Inventario",
                                                              width=100, height=32,
                                                              command=self.cargar_archivo)
        self.boton_subir_inventario.grid(row=2, column=0, sticky="s", padx=10, pady=10)

        # Botón para guardar en la base de datos
        self.boton_guardar_bd = customtkinter.CTkButton(self, text="Guardar en BD",
                                                        width=100, height=32,
                                                        command=self.guardar_en_bd)
        self.boton_guardar_bd.grid(row=2, column=0, sticky="ne", padx=10, pady=10)

        # Botón para volver a la vista principal en la esquina inferior derecha
        self.boton_volver = customtkinter.CTkButton(self, text="Volver", command=volver_func,
                                                    width=100, height=32)
        self.boton_volver.grid(row=2, column=0, sticky="se", padx=10, pady=10)

        # Actualizar el frame
        self.update_idletasks()

    def llamar_menu_opciones(self, elegir):
        """ función que va a manejar las opciones de elegir los inventarios """
        print(f"Option selected: {elegir}")

    def mostrar_panel(self):
        """ Función que muestra el panel """
        label_titulo = customtkinter.CTkLabel(self, text="Bienvenidos a Panel de Inventarios",
                                              font=("Arial", 24), text_color="#021b2b")
        label_titulo.grid(row=0, column=0, pady=20, sticky="n")  
    
    def abrir_agregar_inventario(self):
        """ Función para abrir la ventana de agregar inventario """
        ventana = AgregarInventarioVentana(self)
        ventana.lift()
        ventana.focus_set()

    def cargar_archivo(self):
        """ Función que se encargar de buscar el archivo excel y 
            guardar la información para guardar en la base de datos"""
        ruta_archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        
        if not ruta_archivo:
            return
        
        excel_data_frame = openpyxl.load_workbook(ruta_archivo, data_only=True)  # 'data_only=True' es crucial
        data_frame = excel_data_frame.active
        
        # Se obtiene la información de la hoja de excel
        self.data = []
        
        # Iterar 
        for fila in range(8, data_frame.max_row):
            _fila = []
            
            for col in data_frame.iter_cols(1, data_frame.max_column):
                _fila.append(col[fila].value)
                
            if all(cell is None for cell in _fila):
                break
                
            self.data.append(_fila)
        
        print("Datos leídos del Excel:", self.data)  # Mensaje de depuración
        self.mostrar_datos(self.data)
        
    def mostrar_datos(self, data):
        """ Mostrar los datos del archivo Excel en el Treeview """
        for row in self.treeview.get_children():
            self.treeview.delete(row)
        
        for row in data:
            self.treeview.insert("", "end", values=row)
        
        print("Datos mostrados en el Treeview")  # Mensaje de depuración

    def guardar_en_bd(self):
        """ Función para guardar los datos del archivo Excel en la base de datos MySQL """
        # Conexión a la base de datos
        print(self.data)
        try:
            conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="Mike_090P",
                database="hermanatierra",
                port=3306
            )

            if conn.is_connected():
                print("OK")
                cursor = conn.cursor()

                # Filtrar y limpiar los datos
                filtered_data = [tuple(row[:8]) for row in self.data if len(row) >= 8 and row[0] is not None]
                print("Datos filtrados:", filtered_data)

                if filtered_data:
                    cursor.executemany("""
                        INSERT INTO inventario_utiles_oficinas (id, suministro, presentacion, lote, fv, cantidad, precio, total)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, filtered_data)

                    conn.commit()
                    print(f"{cursor.rowcount} registros insertados.")

        except Error as ex:
            print("Error durante la conexion: {}".format(ex))
        finally:
            if conn.is_connected():
                conn.close()
                print("Conexion cerrada")






